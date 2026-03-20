from pathlib import Path

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation


def write_rows(ws, rows):
    for row in rows:
        ws.append(row)


wb = openpyxl.Workbook()

# 1) Guia de preenchimento
ws_guia = wb.active
ws_guia.title = "guia_preenchimento"
write_rows(
    ws_guia,
    [
        ["aba", "campo", "origem", "obrigatorio", "observacao"],
        ["fornecedores", "base_url/auth_*", "fornecedor", "sim", "Dados do parceiro (TIM/GroupLink/Simhub)"],
        ["endpoints", "endpoint_path/metodo", "fornecedor", "sim", "Conforme documentação da API"],
        ["mapeamento_ids_chassi", "register_chassi_interno", "AcquaX", "sim", "Chassi/register é a chave principal"],
        ["mapeamento_ids_chassi", "meter_id_interno", "AcquaX", "nao", "Opcional: deixar vazio quando não existir ID"],
        ["agendamentos", "horario_local/timezone", "AcquaX", "sim", "Define coletas programadas TIM/GL"],
        ["regras_negocio", "retry/backoff/deduplicacao", "AcquaX", "sim", "Regras de operação e resiliência"],
        ["criterios_aceite", "resultado_esperado", "AcquaX + fornecedor", "sim", "Critérios de homologação"],
    ],
)

# 2) Fornecedores
ws_for = wb.create_sheet("fornecedores")
write_rows(
    ws_for,
    [
        [
            "fornecedor",
            "ambiente",
            "base_url",
            "auth_tipo",
            "auth_detalhes",
            "rate_limit_rps",
            "timeout_s",
            "suporta_webhook",
            "suporta_polling",
            "suporta_agendamento",
            "timezone_padrao",
            "doc_url",
            "contato_tecnico",
        ],
        ["SIMHUB", "producao", "", "", "", "", "", "sim", "sim", "nao", "America/Sao_Paulo", "", ""],
        ["TIM", "producao", "", "", "", "", "", "nao", "sim", "sim", "America/Sao_Paulo", "", ""],
        ["GROUPLINK", "producao", "", "", "", "", "", "nao", "sim", "sim", "America/Sao_Paulo", "", ""],
    ],
)

# 3) Endpoints
ws_end = wb.create_sheet("endpoints")
write_rows(
    ws_end,
    [
        [
            "fornecedor",
            "ambiente",
            "finalidade",
            "metodo",
            "endpoint_path",
            "query_params",
            "body_schema_resumo",
            "response_schema_resumo",
            "usa_paginacao",
            "cursor_ou_pagina",
            "janela_tempo_param",
            "observacoes",
        ],
        ["SIMHUB", "producao", "nivel_tempo_real", "GET", "/reservoir/level", "device_id", "", "", "nao", "", "", ""],
        ["TIM", "producao", "leitura_programada", "GET", "/meters/readings", "from,to,chassi", "", "", "sim", "page", "from,to", ""],
        ["GROUPLINK", "producao", "leitura_programada", "GET", "/iot/readings", "start,end,serial", "", "", "sim", "cursor", "start,end", ""],
    ],
)

# 4) Mapeamento (adaptado para chassi)
ws_map = wb.create_sheet("mapeamento_ids_chassi")
write_rows(
    ws_map,
    [
        [
            "fornecedor",
            "company_id_interno",
            "complex_id_interno",
            "block_id_interno",
            "apartment_id_interno",
            "register_chassi_interno",
            "meter_id_interno_opcional",
            "id_externo_fornecedor",
            "alias_externo",
            "tipo_medidor",
            "ativo",
            "observacao",
        ],
        [
            "TIM",
            "empresa_acquax_01",
            "cond_atlantico_bait",
            "bloco_1",
            "apt_101",
            "CHS-ATB-000145",
            "",
            "tim_device_90871",
            "SERIAL-TIM-90871",
            "agua",
            "sim",
            "Exemplo com chassi como chave",
        ],
        [
            "GROUPLINK",
            "empresa_acquax_01",
            "cond_atlantico_bait",
            "bloco_2",
            "apt_204",
            "CHS-ATB-000146",
            "",
            "gl_sensor_5512",
            "GL-5512",
            "gas",
            "sim",
            "Exemplo com chassi sem meter_id",
        ],
        [
            "SIMHUB",
            "empresa_acquax_01",
            "cond_atlantico_bait",
            "",
            "",
            "CHS-RES-000501",
            "",
            "simhub_tank_501",
            "CAIXA_A_SUPERIOR",
            "nivel_reservatorio",
            "sim",
            "Exemplo para reservatório",
        ],
    ],
)

# 5) Agendamentos
ws_ag = wb.create_sheet("agendamentos")
write_rows(
    ws_ag,
    [
        [
            "fornecedor",
            "company_id_interno",
            "complex_id_interno",
            "register_chassi_interno",
            "frequencia",
            "horario_local",
            "timezone",
            "janela_minutos",
            "ativo",
            "observacao",
        ],
        ["TIM", "empresa_acquax_01", "cond_atlantico_bait", "CHS-ATB-000145", "diaria", "03:00", "America/Sao_Paulo", "15", "sim", ""],
        ["GROUPLINK", "empresa_acquax_01", "cond_atlantico_bait", "CHS-ATB-000146", "diaria", "03:10", "America/Sao_Paulo", "15", "sim", ""],
        ["SIMHUB", "empresa_acquax_01", "cond_atlantico_bait", "CHS-RES-000501", "tempo_real", "", "America/Sao_Paulo", "", "sim", ""],
    ],
)

# 6) Regras de negócio
ws_reg = wb.create_sheet("regras_negocio")
write_rows(
    ws_reg,
    [
        ["chave", "valor", "obrigatorio", "observacao"],
        ["deduplicacao_por", "fornecedor+register_chassi_interno+timestamp", "sim", ""],
        ["atraso_maximo_aceitavel_min", "120", "sim", "Dados fora da janela vão para fila de revisão"],
        ["retry_max_tentativas", "5", "sim", ""],
        ["retry_backoff_segundos", "5,15,30,60,120", "sim", ""],
        ["tratamento_sem_chassi", "rejeitar", "sim", "Sem chassi não integra no fluxo padrão"],
    ],
)

# 7) Critérios de aceite
ws_cri = wb.create_sheet("criterios_aceite")
write_rows(
    ws_cri,
    [
        ["cenario", "entrada", "resultado_esperado", "status"],
        ["TIM leitura válida", "chassi conhecido + leitura", "Leitura salva no medidor correto", "pendente"],
        ["GL leitura duplicada", "mesmo chassi+timestamp", "Ignorar duplicata", "pendente"],
        ["Simhub tempo real", "evento de nível", "Dashboard reflete novo nível", "pendente"],
        ["Sem chassi", "payload sem serial/chassi", "Rejeitado com log e alerta", "pendente"],
    ],
)

# Validações simples
yes_no = DataValidation(type="list", formula1='"sim,nao"', allow_blank=True)
status_dv = DataValidation(type="list", formula1='"pendente,ok,nao_ok"', allow_blank=True)

ws_for.add_data_validation(yes_no)
for col in ["H", "I", "J"]:
    yes_no.add(f"{col}2:{col}2000")

ws_map.add_data_validation(yes_no)
yes_no.add("K2:K2000")

ws_ag.add_data_validation(yes_no)
yes_no.add("I2:I2000")

ws_cri.add_data_validation(status_dv)
status_dv.add("D2:D2000")

# Largura de colunas para legibilidade
for ws in wb.worksheets:
    for col_cells in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(12, max_len + 2), 70)

output = Path("/workspace/docs/integracao_api_template_v2_chassi.xlsx")
output.parent.mkdir(parents=True, exist_ok=True)
wb.save(output)

print(f"Arquivo gerado: {output}")
